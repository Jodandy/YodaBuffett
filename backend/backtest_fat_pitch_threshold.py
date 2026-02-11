#!/usr/bin/env python3
"""
Fat Pitch Threshold Strategy Backtest

Strategy:
- BUY when Fat Pitch score crosses ABOVE 70
- SELL when Fat Pitch score drops BELOW exit threshold (default: 60)

Uses historical dimension scores and price data to calculate returns.
"""

import asyncio
import asyncpg
from datetime import date, timedelta
from typing import Optional
from dataclasses import dataclass
from collections import defaultdict

# Database connection
DATABASE_URL = "postgresql://yodabuffett:password@localhost:5432/yodabuffett"

# Strategy parameters
ENTRY_THRESHOLD = 70  # Buy when score >= this
EXIT_THRESHOLD = 60   # Sell when score < this
HOLD_MINIMUM_DAYS = 30  # Minimum hold period to avoid whipsaws

# Weight profiles
WEIGHT_PROFILES = {
    'garp': {
        'growth': 0.30,           # Strong growth
        'value': 0.25,            # But reasonable valuation
        'profitability': 0.15,    # Profitable
        'earnings_quality': 0.10, # Real earnings
        'returns': 0.10,          # Good returns on capital
        'quality': 0.05,          # Quality business
        'momentum': 0.05,         # Some positive momentum
    },
    'optimal': {
        'value': 0.20,
        'momentum': 0.15,
        'quality': 0.15,
        'profitability': 0.10,
        'growth': 0.10,
        'financial_health': 0.08,
        'returns': 0.07,
        'earnings_quality': 0.05,
        'capital_allocation': 0.05,
        'risk': 0.03,
        'sentiment': 0.02,
    }
}

def build_weight_sql(profile: str = 'garp') -> tuple:
    """Build SQL CASE statements for weighted score calculation."""
    weights = WEIGHT_PROFILES.get(profile, WEIGHT_PROFILES['garp'])

    weight_cases = []
    weight_sum_cases = []

    for dim, weight in weights.items():
        if weight > 0:
            weight_cases.append(f"WHEN dimension_code = '{dim}' THEN score * {weight}")
            weight_sum_cases.append(f"WHEN dimension_code = '{dim}' THEN {weight}")

    weight_sql = " ".join(weight_cases)
    weight_sum_sql = " ".join(weight_sum_cases)

    return weight_sql, weight_sum_sql


@dataclass
class Trade:
    """Represents a completed trade"""
    company_id: str
    symbol: str
    company_name: str
    entry_date: date
    entry_price: float
    entry_score: float
    exit_date: date
    exit_price: float
    exit_score: float
    hold_days: int
    return_pct: float

    @property
    def return_per_day(self) -> float:
        """Return percentage per day held"""
        return self.return_pct / self.hold_days if self.hold_days > 0 else 0.0


@dataclass
class OpenPosition:
    """Represents an open position"""
    company_id: str
    symbol: str
    company_name: str
    entry_date: date
    entry_price: float
    entry_score: float


async def get_historical_scores(conn: asyncpg.Connection, weight_profile: str = 'garp') -> dict:
    """
    Get all historical fat pitch scores with company info.
    Returns dict: company_id -> [(score_date, score, symbol, name), ...]
    """
    # Build weight SQL based on profile
    weight_sql, weight_sum_sql = build_weight_sql(weight_profile)

    query = f"""
    WITH company_scores AS (
        SELECT
            dds.company_id,
            dds.score_date,
            cm.primary_ticker as symbol,
            cm.company_name,
            -- Calculate weighted fat pitch score with proper normalization
            SUM(CASE {weight_sql} ELSE 0 END) /
            NULLIF(SUM(CASE {weight_sum_sql} ELSE 0 END), 0) as fat_pitch_score
        FROM daily_dimension_scores dds
        JOIN company_master cm ON cm.id = dds.company_id
        WHERE cm.primary_ticker IS NOT NULL
        GROUP BY dds.company_id, dds.score_date, cm.primary_ticker, cm.company_name
        HAVING COUNT(DISTINCT dimension_code) >= 5  -- Need enough dimensions
    )
    SELECT company_id, score_date, symbol, company_name, fat_pitch_score
    FROM company_scores
    WHERE fat_pitch_score > 0
    ORDER BY company_id, score_date
    """

    rows = await conn.fetch(query)

    scores_by_company = defaultdict(list)
    for row in rows:
        scores_by_company[row['company_id']].append({
            'date': row['score_date'],
            'score': float(row['fat_pitch_score']),
            'symbol': row['symbol'],
            'name': row['company_name']
        })

    return scores_by_company


async def get_price_on_date(conn: asyncpg.Connection, symbol: str, target_date: date) -> Optional[float]:
    """Get the closing price on or shortly after a given date"""
    query = """
    SELECT close_price, date
    FROM daily_price_data
    WHERE symbol = $1 AND date >= $2
    ORDER BY date ASC
    LIMIT 1
    """
    row = await conn.fetchrow(query, symbol, target_date)
    if row:
        return float(row['close_price'])
    return None


async def get_latest_price(conn: asyncpg.Connection, symbol: str) -> Optional[tuple]:
    """Get the latest price and date for a symbol"""
    query = """
    SELECT close_price, date
    FROM daily_price_data
    WHERE symbol = $1
    ORDER BY date DESC
    LIMIT 1
    """
    row = await conn.fetchrow(query, symbol)
    if row:
        return float(row['close_price']), row['date']
    return None, None


async def run_backtest(
    entry_threshold: float = ENTRY_THRESHOLD,
    exit_threshold: float = EXIT_THRESHOLD,
    min_hold_days: int = HOLD_MINIMUM_DAYS,
    weight_profile: str = 'garp'
):
    """Run the threshold-based backtest"""

    print(f"\n{'='*60}")
    print(f"FAT PITCH THRESHOLD STRATEGY BACKTEST - {weight_profile.upper()} WEIGHTS")
    print(f"{'='*60}")
    print(f"Entry threshold: >= {entry_threshold}")
    print(f"Exit threshold:  < {exit_threshold}")
    print(f"Min hold period: {min_hold_days} days")
    print(f"Weight profile: {weight_profile}")
    print(f"{'='*60}\n")

    conn = await asyncpg.connect(DATABASE_URL)

    try:
        # Get all historical scores
        print(f"Loading historical scores with {weight_profile} weights...")
        scores_by_company = await get_historical_scores(conn, weight_profile)
        print(f"Found {len(scores_by_company)} companies with score history\n")

        # Track completed trades
        completed_trades: list[Trade] = []

        # Process each company
        companies_with_signals = 0

        for company_id, score_history in scores_by_company.items():
            if len(score_history) < 2:
                continue

            symbol = score_history[0]['symbol']
            name = score_history[0]['name']

            position: Optional[OpenPosition] = None
            prev_score = score_history[0]['score']

            for i, data in enumerate(score_history[1:], 1):
                current_date = data['date']
                current_score = data['score']

                # Check for entry signal: score crosses above threshold
                if position is None and prev_score < entry_threshold and current_score >= entry_threshold:
                    # Get entry price
                    entry_price = await get_price_on_date(conn, symbol, current_date)
                    if entry_price:
                        position = OpenPosition(
                            company_id=company_id,
                            symbol=symbol,
                            company_name=name,
                            entry_date=current_date,
                            entry_price=entry_price,
                            entry_score=current_score
                        )
                        companies_with_signals += 1

                # Check for exit signal: score drops below threshold
                elif position is not None:
                    hold_days = (current_date - position.entry_date).days

                    if current_score < exit_threshold and hold_days >= min_hold_days:
                        # Get exit price
                        exit_price = await get_price_on_date(conn, symbol, current_date)
                        if exit_price:
                            return_pct = ((exit_price - position.entry_price) / position.entry_price) * 100

                            trade = Trade(
                                company_id=company_id,
                                symbol=symbol,
                                company_name=name,
                                entry_date=position.entry_date,
                                entry_price=position.entry_price,
                                entry_score=position.entry_score,
                                exit_date=current_date,
                                exit_price=exit_price,
                                exit_score=current_score,
                                hold_days=hold_days,
                                return_pct=return_pct
                            )
                            completed_trades.append(trade)
                            position = None

                prev_score = current_score

            # Close any open position at end of backtest period
            if position is not None:
                # Get the latest price to close the position
                last_score_date = score_history[-1]['date']
                exit_price = await get_price_on_date(conn, symbol, last_score_date)
                if exit_price:
                    hold_days = (last_score_date - position.entry_date).days
                    return_pct = ((exit_price - position.entry_price) / position.entry_price) * 100

                    trade = Trade(
                        company_id=company_id,
                        symbol=symbol,
                        company_name=name,
                        entry_date=position.entry_date,
                        entry_price=position.entry_price,
                        entry_score=position.entry_score,
                        exit_date=last_score_date,
                        exit_price=exit_price,
                        exit_score=score_history[-1]['score'],
                        hold_days=hold_days,
                        return_pct=return_pct
                    )
                    completed_trades.append(trade)

        # Print results
        print(f"\n{'='*60}")
        print("BACKTEST RESULTS")
        print(f"{'='*60}\n")

        if completed_trades:
            # Sort by return
            completed_trades.sort(key=lambda t: t.return_pct, reverse=True)

            returns = [t.return_pct for t in completed_trades]
            avg_return = sum(returns) / len(returns)
            win_rate = len([r for r in returns if r > 0]) / len(returns) * 100
            avg_hold = sum(t.hold_days for t in completed_trades) / len(completed_trades)

            print(f"Completed Trades: {len(completed_trades)}")
            print(f"Win Rate: {win_rate:.1f}%")
            print(f"Average Return: {avg_return:.2f}%")
            print(f"Average Hold Period: {avg_hold:.0f} days")
            print(f"Total Return (sum): {sum(returns):.2f}%")
            print(f"Best Trade: {max(returns):.2f}%")
            print(f"Worst Trade: {min(returns):.2f}%")

            # Show top 10 trades
            print(f"\n{'='*60}")
            print("TOP 10 TRADES")
            print(f"{'='*60}")
            print(f"{'Symbol':<12} {'Entry':<12} {'Exit':<12} {'Days':<6} {'Return':<10} {'Ret/Day':<10} {'Entry Score':<12}")
            print("-" * 80)
            for trade in completed_trades[:10]:
                print(f"{trade.symbol:<12} {str(trade.entry_date):<12} {str(trade.exit_date):<12} "
                      f"{trade.hold_days:<6} {trade.return_pct:>+7.2f}%   {trade.return_per_day:>+6.3f}%   {trade.entry_score:.1f}")

            # Show worst 10 trades
            print(f"\n{'='*60}")
            print("WORST 10 TRADES")
            print(f"{'='*60}")
            print(f"{'Symbol':<12} {'Entry':<12} {'Exit':<12} {'Days':<6} {'Return':<10} {'Ret/Day':<10} {'Exit Score':<12}")
            print("-" * 80)
            for trade in completed_trades[-10:]:
                print(f"{trade.symbol:<12} {str(trade.entry_date):<12} {str(trade.exit_date):<12} "
                      f"{trade.hold_days:<6} {trade.return_pct:>+7.2f}%   {trade.return_per_day:>+6.3f}%   {trade.exit_score:.1f}")
        else:
            print("No completed trades found.")

        # Summary by year
        if completed_trades:
            print(f"\n{'='*60}")
            print("RETURNS BY YEAR")
            print(f"{'='*60}")

            by_year = defaultdict(list)
            for trade in completed_trades:
                year = trade.exit_date.year
                by_year[year].append(trade.return_pct)

            for year in sorted(by_year.keys()):
                returns = by_year[year]
                avg = sum(returns) / len(returns)
                wins = len([r for r in returns if r > 0])
                print(f"{year}: {len(returns):>3} trades, {wins/len(returns)*100:>5.1f}% win rate, {avg:>+7.2f}% avg return")

    finally:
        await conn.close()


async def test_thresholds(weight_profile: str = 'garp'):
    """Test different threshold combinations - FULL MATRIX"""
    print("\n" + "="*70)
    print(f"THRESHOLD SENSITIVITY ANALYSIS - {weight_profile.upper()} WEIGHTS")
    print("="*70 + "\n")

    results = []

    # Full matrix: coarse at low end, fine granularity at top end (where alpha is)
    # Low end: 40, 45, 50, 55 (steps of 5)
    # High end: 58, 60, 62, 64, 66, 68, 70, 72, 74, 76, 78, 80 (steps of 2)
    thresholds = [40, 45, 50, 55] + list(range(58, 82, 2))

    total_combos = sum(1 for e in thresholds for x in thresholds if x < e)
    print(f"Testing {total_combos} entry/exit combinations...")
    print(f"Entry range: {min(thresholds)}-{max(thresholds)}")
    print(f"Exit range: {min(thresholds)}-{max(thresholds)}")
    print(f"Weight profile: {weight_profile}")
    print()

    for entry in thresholds:
        for exit in thresholds:
            if exit >= entry:
                continue

            conn = await asyncpg.connect(DATABASE_URL)
            try:
                scores_by_company = await get_historical_scores(conn, weight_profile)

                trades = []
                for company_id, score_history in scores_by_company.items():
                    if len(score_history) < 2:
                        continue

                    symbol = score_history[0]['symbol']
                    position = None
                    prev_score = score_history[0]['score']

                    for data in score_history[1:]:
                        current_date = data['date']
                        current_score = data['score']

                        if position is None and prev_score < entry and current_score >= entry:
                            entry_price = await get_price_on_date(conn, symbol, current_date)
                            if entry_price:
                                position = {
                                    'entry_date': current_date,
                                    'entry_price': entry_price,
                                    'symbol': symbol
                                }

                        elif position is not None:
                            hold_days = (current_date - position['entry_date']).days
                            if current_score < exit and hold_days >= 30:
                                exit_price = await get_price_on_date(conn, symbol, current_date)
                                if exit_price:
                                    ret = ((exit_price - position['entry_price']) / position['entry_price']) * 100
                                    trades.append(ret)
                                    position = None

                        prev_score = current_score

                    # Close any open position at end of backtest period
                    if position is not None:
                        last_score_date = score_history[-1]['date']
                        exit_price = await get_price_on_date(conn, symbol, last_score_date)
                        if exit_price:
                            ret = ((exit_price - position['entry_price']) / position['entry_price']) * 100
                            trades.append(ret)

                if trades:
                    avg_ret = sum(trades) / len(trades)
                    win_rate = len([r for r in trades if r > 0]) / len(trades) * 100
                    results.append({
                        'entry': entry,
                        'exit': exit,
                        'trades': len(trades),
                        'avg_return': avg_ret,
                        'win_rate': win_rate,
                        'total_return': sum(trades)
                    })

            finally:
                await conn.close()

    # Print results sorted by average return
    print("\n" + "="*70)
    print("RESULTS SORTED BY AVERAGE RETURN")
    print("="*70)
    print(f"{'Entry':<8} {'Exit':<8} {'Trades':<8} {'Win Rate':<10} {'Avg Ret':<12} {'Total Ret':<12}")
    print("-" * 70)

    for r in sorted(results, key=lambda x: x['avg_return'], reverse=True):
        print(f"{r['entry']:<8} {r['exit']:<8} {r['trades']:<8} {r['win_rate']:>5.1f}%     "
              f"{r['avg_return']:>+8.2f}%    {r['total_return']:>+10.1f}%")

    # Print matrix view for average return
    print("\n" + "="*70)
    print("MATRIX VIEW: AVERAGE RETURN BY ENTRY/EXIT")
    print("="*70)
    print("(Rows = Entry threshold, Columns = Exit threshold)")
    print()

    # Get unique thresholds
    entries = sorted(set(r['entry'] for r in results))
    exits = sorted(set(r['exit'] for r in results))

    # Header
    print(f"{'Entry↓ Exit→':<12}", end="")
    for exit in exits:
        print(f"{exit:>8}", end="")
    print()
    print("-" * (12 + 8 * len(exits)))

    # Matrix
    for entry in entries:
        print(f"{entry:<12}", end="")
        for exit in exits:
            if exit >= entry:
                print(f"{'---':>8}", end="")
            else:
                match = next((r for r in results if r['entry'] == entry and r['exit'] == exit), None)
                if match and match['trades'] >= 5:  # Only show if >= 5 trades
                    print(f"{match['avg_return']:>+7.1f}%", end="")
                elif match:
                    print(f"{'(n<5)':>8}", end="")
                else:
                    print(f"{'N/A':>8}", end="")
        print()

    # Print matrix view for win rate
    print("\n" + "="*70)
    print("MATRIX VIEW: WIN RATE BY ENTRY/EXIT")
    print("="*70)
    print()

    print(f"{'Entry↓ Exit→':<12}", end="")
    for exit in exits:
        print(f"{exit:>8}", end="")
    print()
    print("-" * (12 + 8 * len(exits)))

    for entry in entries:
        print(f"{entry:<12}", end="")
        for exit in exits:
            if exit >= entry:
                print(f"{'---':>8}", end="")
            else:
                match = next((r for r in results if r['entry'] == entry and r['exit'] == exit), None)
                if match and match['trades'] >= 5:
                    print(f"{match['win_rate']:>6.1f}%", end=" ")
                elif match:
                    print(f"{'(n<5)':>8}", end="")
                else:
                    print(f"{'N/A':>8}", end="")
        print()

    # Print matrix view for trade count
    print("\n" + "="*70)
    print("MATRIX VIEW: NUMBER OF TRADES BY ENTRY/EXIT")
    print("="*70)
    print()

    print(f"{'Entry↓ Exit→':<12}", end="")
    for exit in exits:
        print(f"{exit:>8}", end="")
    print()
    print("-" * (12 + 8 * len(exits)))

    for entry in entries:
        print(f"{entry:<12}", end="")
        for exit in exits:
            if exit >= entry:
                print(f"{'---':>8}", end="")
            else:
                match = next((r for r in results if r['entry'] == entry and r['exit'] == exit), None)
                if match:
                    print(f"{match['trades']:>8}", end="")
                else:
                    print(f"{'0':>8}", end="")
        print()

    # Summary insights
    print("\n" + "="*70)
    print("KEY INSIGHTS")
    print("="*70)

    # Best by avg return with min 10 trades
    valid = [r for r in results if r['trades'] >= 10]
    if valid:
        best = max(valid, key=lambda x: x['avg_return'])
        print(f"\nBest avg return (min 10 trades):")
        print(f"  Entry {best['entry']} / Exit {best['exit']}: {best['avg_return']:+.2f}% avg, "
              f"{best['win_rate']:.1f}% win rate, {best['trades']} trades")

    # Best by win rate with min 10 trades
    if valid:
        best_wr = max(valid, key=lambda x: x['win_rate'])
        print(f"\nBest win rate (min 10 trades):")
        print(f"  Entry {best_wr['entry']} / Exit {best_wr['exit']}: {best_wr['win_rate']:.1f}% win rate, "
              f"{best_wr['avg_return']:+.2f}% avg, {best_wr['trades']} trades")

    # Most trades with positive avg
    positive = [r for r in results if r['avg_return'] > 0]
    if positive:
        most_trades = max(positive, key=lambda x: x['trades'])
        print(f"\nMost trades with positive avg return:")
        print(f"  Entry {most_trades['entry']} / Exit {most_trades['exit']}: {most_trades['trades']} trades, "
              f"{most_trades['avg_return']:+.2f}% avg, {most_trades['win_rate']:.1f}% win rate")


async def export_all_trades_to_excel(output_file: str = "fat_pitch_threshold_trades.xlsx", filter_combos: bool = True):
    """Export all trades for all threshold combinations to Excel

    Args:
        output_file: Output Excel filename
        filter_combos: If True, only export best + worst combos (not all 105+)
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Please install pandas and openpyxl: pip install pandas openpyxl")
        return

    print("\n" + "="*70)
    print("EXPORTING TRADES TO EXCEL")
    print("="*70 + "\n")

    # Outlier thresholds - filter likely bad data (reverse splits, etc.)
    MAX_RETURN = 200.0  # Filter trades > 200%
    MIN_RETURN = -80.0  # Filter trades < -80%

    # Selected threshold combos to include (best + worst for comparison)
    # These are chosen based on prior analysis with outliers filtered
    SELECTED_COMBOS = [
        # Higher thresholds (more selective entry)
        (80, 70), (78, 68), (76, 66), (74, 64), (72, 62), (70, 60), (68, 58),
        # Mid thresholds
        (66, 56), (64, 54), (62, 52), (60, 50),
        # Lower thresholds (wider net)
        (58, 48), (55, 45), (50, 40),
    ] if filter_combos else None

    print(f"Outlier filter: Excluding returns > {MAX_RETURN}% or < {MIN_RETURN}%")
    if filter_combos:
        print(f"Selected combos: {len(SELECTED_COMBOS)} combinations")
    else:
        print("Including ALL combinations")
    print()

    conn = await asyncpg.connect(DATABASE_URL)

    try:
        # Get all historical scores
        print("Loading historical scores...")
        scores_by_company = await get_historical_scores(conn)
        print(f"Found {len(scores_by_company)} companies with score history\n")

        # Thresholds to test
        if filter_combos:
            # Use selected combos
            threshold_pairs = SELECTED_COMBOS
        else:
            # Full matrix
            thresholds = [40, 45, 50, 55] + list(range(58, 82, 2))
            threshold_pairs = [(e, x) for e in thresholds for x in thresholds if x < e]

        # Store all results
        all_results = []
        all_trades_by_combo = {}
        outliers_filtered = 0

        total_combos = len(threshold_pairs)
        print(f"Processing {total_combos} combinations...")

        combo_count = 0
        for entry, exit in threshold_pairs:
            combo_count += 1
            combo_key = f"E{entry}_X{exit}"
            trades = []

            for company_id, score_history in scores_by_company.items():
                if len(score_history) < 2:
                    continue

                symbol = score_history[0]['symbol']
                name = score_history[0]['name']

                position = None
                prev_score = score_history[0]['score']

                for data in score_history[1:]:
                    current_date = data['date']
                    current_score = data['score']

                    # Entry signal
                    if position is None and prev_score < entry and current_score >= entry:
                        entry_price = await get_price_on_date(conn, symbol, current_date)
                        if entry_price:
                            position = {
                                'symbol': symbol,
                                'name': name,
                                'entry_date': current_date,
                                'entry_price': entry_price,
                                'entry_score': current_score
                            }

                    # Exit signal
                    elif position is not None:
                        hold_days = (current_date - position['entry_date']).days
                        if current_score < exit and hold_days >= 30:
                            exit_price = await get_price_on_date(conn, symbol, current_date)
                            if exit_price:
                                ret = ((exit_price - position['entry_price']) / position['entry_price']) * 100
                                # Filter outliers (likely bad data from splits)
                                if ret > MAX_RETURN or ret < MIN_RETURN:
                                    outliers_filtered += 1
                                    position = None
                                    prev_score = current_score
                                    continue
                                trades.append({
                                    'Symbol': position['symbol'],
                                    'Company': position['name'],
                                    'Entry Date': position['entry_date'],
                                    'Entry Price': position['entry_price'],
                                    'Entry Score': position['entry_score'],
                                    'Exit Date': current_date,
                                    'Exit Price': exit_price,
                                    'Exit Score': current_score,
                                    'Hold Days': hold_days,
                                    'Return %': ret,
                                    'Return/Day %': ret / hold_days if hold_days > 0 else 0
                                })
                                position = None

                    prev_score = current_score

                # Close any open position at end of backtest period
                if position is not None:
                    last_score_date = score_history[-1]['date']
                    exit_price = await get_price_on_date(conn, symbol, last_score_date)
                    if exit_price:
                        hold_days = (last_score_date - position['entry_date']).days
                        ret = ((exit_price - position['entry_price']) / position['entry_price']) * 100
                        # Filter outliers (likely bad data from splits)
                        if ret <= MAX_RETURN and ret >= MIN_RETURN:
                            trades.append({
                                'Symbol': position['symbol'],
                                'Company': position['name'],
                                'Entry Date': position['entry_date'],
                                'Entry Price': position['entry_price'],
                                'Entry Score': position['entry_score'],
                                'Exit Date': last_score_date,
                                'Exit Price': exit_price,
                                'Exit Score': score_history[-1]['score'],
                                'Hold Days': hold_days,
                                'Return %': ret,
                                'Return/Day %': ret / hold_days if hold_days > 0 else 0
                            })
                        else:
                            outliers_filtered += 1

            if trades:
                all_trades_by_combo[combo_key] = trades
                returns = [t['Return %'] for t in trades]
                all_results.append({
                    'Entry': entry,
                    'Exit': exit,
                    'Trades': len(trades),
                    'Win Rate %': len([r for r in returns if r > 0]) / len(returns) * 100,
                    'Avg Return %': sum(returns) / len(returns),
                    'Total Return %': sum(returns),
                    'Best Trade %': max(returns),
                    'Worst Trade %': min(returns)
                })

            if combo_count % 5 == 0:
                print(f"  Processed {combo_count}/{total_combos} combinations...")

        print(f"\nOutliers filtered: {outliers_filtered} trades (>{MAX_RETURN}% or <{MIN_RETURN}%)")
        print(f"Creating Excel file: {output_file}")

        # Create workbook
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Write summary header
        summary_headers = ['Entry', 'Exit', 'Trades', 'Win Rate %', 'Avg Return %', 'Total Return %', 'Best Trade %', 'Worst Trade %']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Sort by avg return descending
        all_results.sort(key=lambda x: x['Avg Return %'], reverse=True)

        for row_idx, result in enumerate(all_results, 2):
            ws_summary.cell(row=row_idx, column=1, value=result['Entry'])
            ws_summary.cell(row=row_idx, column=2, value=result['Exit'])
            ws_summary.cell(row=row_idx, column=3, value=result['Trades'])
            ws_summary.cell(row=row_idx, column=4, value=round(result['Win Rate %'], 1))
            ws_summary.cell(row=row_idx, column=5, value=round(result['Avg Return %'], 2))
            ws_summary.cell(row=row_idx, column=6, value=round(result['Total Return %'], 1))
            ws_summary.cell(row=row_idx, column=7, value=round(result['Best Trade %'], 2))
            ws_summary.cell(row=row_idx, column=8, value=round(result['Worst Trade %'], 2))

            # Color code avg return
            avg_cell = ws_summary.cell(row=row_idx, column=5)
            if result['Avg Return %'] > 0:
                avg_cell.fill = green_fill
            else:
                avg_cell.fill = red_fill

            for col in range(1, 9):
                ws_summary.cell(row=row_idx, column=col).border = thin_border

        # Adjust column widths
        for col in range(1, 9):
            ws_summary.column_dimensions[chr(64 + col)].width = 15

        # Create a sheet for each combination with trades
        print(f"Creating {len(all_trades_by_combo)} trade sheets...")

        # Sort combos by avg return for tab ordering
        combo_order = sorted(all_trades_by_combo.keys(),
                           key=lambda k: next((r['Avg Return %'] for r in all_results
                                              if f"E{r['Entry']}_X{r['Exit']}" == k), -999),
                           reverse=True)

        for combo_key in combo_order:
            trades = all_trades_by_combo[combo_key]
            if len(trades) < 1:
                continue

            # Create sheet (Excel tab names max 31 chars)
            ws = wb.create_sheet(title=combo_key[:31])

            # Headers
            trade_headers = ['Symbol', 'Company', 'Entry Date', 'Entry Price', 'Entry Score',
                           'Exit Date', 'Exit Price', 'Exit Score', 'Hold Days', 'Return %', 'Return/Day %']

            for col, header in enumerate(trade_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Sort trades by return descending
            trades.sort(key=lambda x: x['Return %'], reverse=True)

            # Write trades
            for row_idx, trade in enumerate(trades, 2):
                ws.cell(row=row_idx, column=1, value=trade['Symbol'])
                ws.cell(row=row_idx, column=2, value=trade['Company'][:30] if trade['Company'] else '')
                ws.cell(row=row_idx, column=3, value=str(trade['Entry Date']))
                ws.cell(row=row_idx, column=4, value=round(trade['Entry Price'], 2))
                ws.cell(row=row_idx, column=5, value=round(trade['Entry Score'], 1))
                ws.cell(row=row_idx, column=6, value=str(trade['Exit Date']))
                ws.cell(row=row_idx, column=7, value=round(trade['Exit Price'], 2))
                ws.cell(row=row_idx, column=8, value=round(trade['Exit Score'], 1))
                ws.cell(row=row_idx, column=9, value=trade['Hold Days'])

                ret_cell = ws.cell(row=row_idx, column=10, value=round(trade['Return %'], 2))
                if trade['Return %'] > 0:
                    ret_cell.fill = green_fill
                else:
                    ret_cell.fill = red_fill

                ret_day_cell = ws.cell(row=row_idx, column=11, value=round(trade['Return/Day %'], 4))
                if trade['Return/Day %'] > 0:
                    ret_day_cell.fill = green_fill
                else:
                    ret_day_cell.fill = red_fill

                for col in range(1, 12):
                    ws.cell(row=row_idx, column=col).border = thin_border

            # Adjust column widths
            col_widths = [12, 25, 12, 12, 12, 12, 12, 12, 10, 12, 12]
            for col, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width

        # Save
        wb.save(output_file)
        print(f"\n✅ Excel file saved: {output_file}")
        print(f"   - Summary tab with all {len(all_results)} combinations")
        print(f"   - {len(all_trades_by_combo)} individual trade tabs")
        print(f"   - Tabs ordered by average return (best first)")

    finally:
        await conn.close()


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Fat Pitch Threshold Strategy Backtest")
    parser.add_argument("--entry", type=float, default=70, help="Entry threshold (buy when score >= this)")
    parser.add_argument("--exit", type=float, default=60, help="Exit threshold (sell when score < this)")
    parser.add_argument("--min-hold", type=int, default=30, help="Minimum hold period in days")
    parser.add_argument("--weights", type=str, default='garp', choices=['garp', 'optimal'],
                        help="Weight profile to use (default: garp)")
    parser.add_argument("--sensitivity", action="store_true", help="Run threshold sensitivity analysis")
    parser.add_argument("--export", action="store_true", help="Export trades to Excel")
    parser.add_argument("--all-combos", action="store_true", help="Export ALL threshold combos (default: filtered best/worst)")
    parser.add_argument("--output", type=str, default="fat_pitch_threshold_trades.xlsx", help="Excel output filename")

    args = parser.parse_args()

    if args.export:
        asyncio.run(export_all_trades_to_excel(args.output, filter_combos=not args.all_combos))
    elif args.sensitivity:
        asyncio.run(test_thresholds(args.weights))
    else:
        asyncio.run(run_backtest(args.entry, args.exit, args.min_hold, args.weights))
